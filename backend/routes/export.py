"""
routes/export.py
=================
Data export endpoints — download filtered data in multiple formats.

GET /api/export/csv       — download as CSV
GET /api/export/xlsx      — download as Excel (.xlsx)
GET /api/export/json      — download as JSON
GET /api/export/tsv       — download as TSV
GET /api/export/summary   — KPI summary report as CSV
"""

import io
import json
import csv
from datetime import datetime

import pandas as pd
from flask import Blueprint, request, jsonify, send_file, Response

from utils.data_processor import (
    get_session_df, get_latest_session_id,
    apply_filters, compute_kpis,
    by_category, by_region, by_rep,
)

export_bp = Blueprint("export", __name__)

EXPORT_COLUMNS = ["id", "date", "category", "product", "region",
                  "rep", "units", "revenue", "cost", "profit", "customer"]


def _resolve_df():
    session_id = request.args.get("session_id") or get_latest_session_id()
    if not session_id:
        return None, jsonify({"error": "No data loaded. Upload a file first."}), 400
    df = get_session_df(session_id)
    if df is None:
        return None, jsonify({"error": f"Session '{session_id}' not found"}), 404
    return df, None, None


def _filter_and_select(df):
    params = {
        "region":    request.args.get("region", ""),
        "category":  request.args.get("category", ""),
        "rep":       request.args.get("rep", ""),
        "date_from": request.args.get("date_from", ""),
        "date_to":   request.args.get("date_to", ""),
    }
    filtered = apply_filters(df, params)

    # Only include columns that exist
    cols = [c for c in EXPORT_COLUMNS if c in filtered.columns]
    out  = filtered[cols].copy()
    if "date" in out.columns:
        out["date"] = out["date"].dt.strftime("%Y-%m-%d").where(out["date"].notna(), "")
    return out


def _ts():
    return datetime.now().strftime("%Y%m%d_%H%M%S")


# ── CSV ────────────────────────────────────────────────────────────────────────
@export_bp.route("/csv", methods=["GET"])
def export_csv():
    """Download filtered data as CSV."""
    df, err, code = _resolve_df()
    if err:
        return err, code

    out = _filter_and_select(df)
    buf = io.StringIO()
    out.to_csv(buf, index=False)
    buf.seek(0)

    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=sales_export_{_ts()}.csv"},
    )


# ── XLSX ───────────────────────────────────────────────────────────────────────
@export_bp.route("/xlsx", methods=["GET"])
def export_xlsx():
    """Download filtered data as a formatted Excel workbook."""
    df, err, code = _resolve_df()
    if err:
        return err, code

    out = _filter_and_select(df)
    buf = io.BytesIO()

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        out.to_excel(writer, sheet_name="Sales Data", index=False)
        wb = writer.book
        ws = writer.sheets["Sales Data"]

        # Style header row
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = Alignment(horizontal="center")

        # Auto-width columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"sales_export_{_ts()}.xlsx",
    )


# ── JSON ───────────────────────────────────────────────────────────────────────
@export_bp.route("/json", methods=["GET"])
def export_json():
    """Download filtered data as JSON."""
    df, err, code = _resolve_df()
    if err:
        return err, code

    out = _filter_and_select(df)
    payload = {
        "exported_at": datetime.now().isoformat(),
        "records":     len(out),
        "data":        out.to_dict(orient="records"),
    }

    return Response(
        json.dumps(payload, indent=2, default=str),
        mimetype="application/json",
        headers={"Content-Disposition": f"attachment; filename=sales_export_{_ts()}.json"},
    )


# ── TSV ────────────────────────────────────────────────────────────────────────
@export_bp.route("/tsv", methods=["GET"])
def export_tsv():
    """Download filtered data as TSV (tab-separated values)."""
    df, err, code = _resolve_df()
    if err:
        return err, code

    out = _filter_and_select(df)
    buf = io.StringIO()
    out.to_csv(buf, index=False, sep="\t")
    buf.seek(0)

    return Response(
        buf.getvalue(),
        mimetype="text/tab-separated-values",
        headers={"Content-Disposition": f"attachment; filename=sales_export_{_ts()}.tsv"},
    )


# ── Summary Report ─────────────────────────────────────────────────────────────
@export_bp.route("/summary", methods=["GET"])
def export_summary():
    """
    Download a KPI + breakdown summary report as CSV.
    Includes: overall KPIs, revenue by category, revenue by region, rep performance.
    """
    df, err, code = _resolve_df()
    if err:
        return err, code

    params = {
        "region":    request.args.get("region", ""),
        "category":  request.args.get("category", ""),
        "rep":       request.args.get("rep", ""),
        "date_from": request.args.get("date_from", ""),
        "date_to":   request.args.get("date_to", ""),
    }
    filtered = apply_filters(df, params)
    kpis     = compute_kpis(filtered)
    cats     = by_category(filtered)
    regs     = by_region(filtered)
    reps     = by_rep(filtered)

    buf  = io.StringIO()
    w    = csv.writer(buf)
    ts   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    w.writerow(["SALESIQ BUSINESS INTELLIGENCE SUMMARY REPORT"])
    w.writerow([f"Generated: {ts}"])
    w.writerow([f"Records analysed: {len(filtered)}"])
    if any(params.values()):
        active = {k: v for k, v in params.items() if v}
        w.writerow([f"Filters: {active}"])
    w.writerow([])

    w.writerow(["── KEY METRICS ──"])
    for k, v in kpis.items():
        if k != "date_range":
            w.writerow([k.replace("_", " ").title(), v])
    w.writerow([])

    w.writerow(["── REVENUE BY CATEGORY ──"])
    w.writerow(["Category", "Revenue ($)", "Profit ($)", "Margin (%)", "Orders"])
    for r in cats:
        w.writerow([r["category"], r["revenue"], r["profit"], r["margin_pct"], r["orders"]])
    w.writerow([])

    w.writerow(["── REVENUE BY REGION ──"])
    w.writerow(["Region", "Revenue ($)", "Profit ($)", "% of Total", "Orders"])
    for r in regs:
        w.writerow([r["region"], r["revenue"], r["profit"], r["pct_of_total"], r["orders"]])
    w.writerow([])

    w.writerow(["── SALES REP PERFORMANCE ──"])
    w.writerow(["Rep", "Revenue ($)", "Profit ($)", "Margin (%)", "Orders"])
    for r in reps:
        w.writerow([r["rep"], r["revenue"], r["profit"], r["margin_pct"], r["orders"]])

    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=summary_report_{_ts()}.csv"},
    )
