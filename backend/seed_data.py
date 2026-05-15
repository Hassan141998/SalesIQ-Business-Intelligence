"""
seed_data.py
=============
Generate realistic sample data files for testing and demo.
Run with: python seed_data.py

Produces:
  sample-data/sales_data.csv        — 120 records
  sample-data/sales_data.xlsx       — formatted Excel
  sample-data/sales_data.json       — JSON format
  sample-data/sales_data_small.tsv  — 20 records TSV
"""

import os
import json
import random
import pandas as pd
from datetime import date, timedelta
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

random.seed(42)

# ── Config ─────────────────────────────────────────────────────────────────────
OUT_DIR = os.path.join(os.path.dirname(__file__), "..", "sample-data")
os.makedirs(OUT_DIR, exist_ok=True)

CATEGORIES = {
    "Electronics":     [("Laptop Pro", 1299.99), ("Smart Watch", 299.99), ("Tablet X1", 599.99),
                        ("Wireless Headphones", 149.99), ("Gaming Console", 499.99),
                        ("Smart TV 55\"", 799.99), ("Bluetooth Speaker", 99.99)],
    "Clothing":        [("Winter Jacket", 89.99), ("Running Shoes", 79.99), ("Denim Jeans", 69.99),
                        ("Hoodie", 54.99), ("Yoga Pants", 49.99), ("Formal Suit", 299.99)],
    "Food & Beverage": [("Organic Coffee Set", 34.99), ("Protein Powder", 54.99),
                        ("Premium Tea Set", 24.99), ("Vitamin Bundle", 49.99),
                        ("Snack Box", 29.99)],
    "Home & Garden":   [("Smart Thermostat", 249.99), ("Robot Vacuum", 449.99),
                        ("Air Purifier", 199.99), ("Standing Desk", 499.99),
                        ("Coffee Table", 249.99)],
}

REGIONS  = ["North", "South", "East", "West"]
REPS     = ["Alice Johnson", "Bob Smith", "Carol White", "David Lee"]
CUSTOMERS= ["TechCorp", "FashionHub", "SoundWorld", "SmartHome Inc",
            "SportZone", "CafeWorld", "WearTech", "CleanHome", "FitLife"]

START_DATE = date(2024, 1, 1)
END_DATE   = date(2024, 12, 31)


def random_date() -> date:
    delta = (END_DATE - START_DATE).days
    return START_DATE + timedelta(days=random.randint(0, delta))


def cost_ratio(cat: str) -> float:
    """Approximate COGS ratio by category."""
    return {"Electronics": 0.60, "Clothing": 0.45,
            "Food & Beverage": 0.50, "Home & Garden": 0.52}.get(cat, 0.55)


def build_records(n: int = 120) -> list[dict]:
    records = []
    for i in range(n):
        cat  = random.choice(list(CATEGORIES))
        prod, base_price = random.choice(CATEGORIES[cat])
        units = random.randint(1, 25)
        price = round(base_price * random.uniform(0.95, 1.05), 2)
        rev   = round(units * price, 2)
        cost  = round(rev * cost_ratio(cat) * random.uniform(0.95, 1.05), 2)
        records.append({
            "Order ID":  1001 + i,
            "Date":      random_date().isoformat(),
            "Category":  cat,
            "Product":   prod,
            "Region":    random.choice(REGIONS),
            "Sales Rep": random.choice(REPS),
            "Units":     units,
            "Unit Price":price,
            "Revenue":   rev,
            "Cost":      cost,
            "Profit":    round(rev - cost, 2),
            "Customer":  random.choice(CUSTOMERS),
        })
    return sorted(records, key=lambda r: r["Date"])


# ── Build ──────────────────────────────────────────────────────────────────────
records = build_records(120)
df      = pd.DataFrame(records)

# CSV
csv_path = os.path.join(OUT_DIR, "sales_data.csv")
df.to_csv(csv_path, index=False)
print(f"✅ CSV  → {csv_path}  ({len(df)} rows)")

# TSV (small)
tsv_path = os.path.join(OUT_DIR, "sales_data_small.tsv")
df.head(20).to_csv(tsv_path, index=False, sep="\t")
print(f"✅ TSV  → {tsv_path}  (20 rows)")

# JSON
json_path = os.path.join(OUT_DIR, "sales_data.json")
payload   = {
    "metadata": {
        "source":      "SalesIQ Seed Data",
        "generated":   date.today().isoformat(),
        "records":     len(records),
        "description": "Sample JSON format for API / system integrations",
    },
    "sales": records[:10],   # keep JSON file small for demo
}
with open(json_path, "w") as f:
    json.dump(payload, f, indent=2)
print(f"✅ JSON → {json_path}  (10 records)")

# XLSX — styled
xlsx_path = os.path.join(OUT_DIR, "sales_data.xlsx")

with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Sales Data", index=False)
    wb = writer.book
    ws = writer.sheets["Sales Data"]

    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    alt_fill    = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")

    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx in range(2, len(df) + 2):
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            cell.alignment = Alignment(horizontal="center")

    col_widths = [10, 12, 16, 22, 10, 16, 8, 11, 12, 12, 12, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 22

print(f"✅ XLSX → {xlsx_path}  ({len(df)} rows, styled)")
print("\n🎉 All sample data files generated successfully!")
